import requests
import bs4
import openpyxl

# Get Data from site 
res = requests.get('https://news.ycombinator.com/news')
res.raise_for_status()
hacker_news = bs4.BeautifulSoup(res.text, 'html.parser')
body = hacker_news.find_all('span', attrs={"class":"titleline"}, limit=5)

for i in body:
    print(i.text, end=' ')
    print("\n" + i.find('a').get('href'),"\n\n")

# Create an active Workbook to write data
wb = openpyxl.Workbook()
wb.create_sheet(index=0, title='Hacker News')
sheet = wb['Hacker News']

for j in range(0, len(body)):
    cell_data = body[j].text + "\n" + body[j].find('a').get('href')
    sheet.cell(row=j+1, column=1).value = cell_data


wb.save('HackerNews_URLs.xlsx')